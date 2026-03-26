import os
import re
import tempfile
from io import BytesIO

import pandas as pd
import pytesseract
import streamlit as st
from PIL import Image, ImageOps
from pdf2image import convert_from_path
from openpyxl import load_workbook


st.set_page_config(page_title="PDF OCR Tracker Updater", layout="wide")
st.title("PDF OCR Tracker Updater")


# -----------------------------
# Helpers
# -----------------------------
def normalize_text(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.startswith("'"):
        value = value[1:]
    value = value.replace("\xa0", " ")
    value = " ".join(value.split())
    return value.upper()


def normalize_header(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def normalize_part_number(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.startswith("'"):
        value = value[1:]
    value = value.replace("\xa0", " ")
    value = " ".join(value.split())
    return value.upper()


def combine_unique_values(series):
    values = []
    for v in series:
        if pd.isna(v):
            continue
        v = str(v).strip()
        if v != "":
            values.append(v)

    unique_values = sorted(set(values), key=lambda x: (len(x), x))
    return ", ".join(unique_values)


def merge_comma_separated(existing_value, new_value):
    existing_items = []
    new_items = []

    if existing_value not in [None, ""]:
        existing_items = [x.strip() for x in str(existing_value).split(",") if x.strip()]

    if new_value not in [None, ""]:
        new_items = [x.strip() for x in str(new_value).split(",") if x.strip()]

    merged = sorted(set(existing_items + new_items), key=lambda x: (len(x), x))
    return ", ".join(merged)


def normalize_quantity_text(qty_raw):
    qty_raw = qty_raw.strip()

    if "," in qty_raw and "." in qty_raw:
        if qty_raw.rfind(",") > qty_raw.rfind("."):
            qty_clean = qty_raw.replace(".", "").replace(",", ".")
        else:
            qty_clean = qty_raw.replace(",", "")
    elif "," in qty_raw:
        parts = qty_raw.split(",")
        if len(parts[-1]) == 3 and len(parts) > 1:
            qty_clean = qty_raw.replace(",", "")
        else:
            qty_clean = qty_raw.replace(",", ".")
    elif "." in qty_raw:
        parts = qty_raw.split(".")
        if len(parts[-1]) == 3 and len(parts) > 1:
            qty_clean = qty_raw.replace(".", "")
        else:
            qty_clean = qty_raw
    else:
        qty_clean = qty_raw

    return int(float(qty_clean))


def extract_document_number(text):
    patterns = [
        r"Truck\s*No\.?\s*[:#]?\s*([A-Z0-9\-\/]+)",
        r"Truck\s*[:#]?\s*([A-Z0-9\-\/]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()

    return ""


def load_pages_from_upload(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()

    if ext == ".pdf":
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name

        try:
            pages = convert_from_path(tmp_path, dpi=300)
            return [page.convert("L") for page in pages]
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    else:
        img = Image.open(uploaded_file).convert("L")
        return [img]


def stitch_pages_vertically(pages):
    widths, heights = zip(*(p.size for p in pages))
    max_width = max(widths)
    total_height = sum(heights)

    stitched = Image.new("L", (max_width, total_height), color=255)

    y_offset = 0
    for page in pages:
        stitched.paste(page, (0, y_offset))
        y_offset += page.size[1]

    return stitched


def crop_table_region(img):
    width, height = img.size
    top_crop = int(height * 0.05)
    bottom_crop = int(height * 1.00)
    cropped = img.crop((0, top_crop, width, bottom_crop))
    return ImageOps.autocontrast(cropped)


def process_delivery_files(delivery_files):
    all_items = []
    preview_images = []

    for uploaded_file in delivery_files:
        pages = load_pages_from_upload(uploaded_file)
        stitched_img = stitch_pages_vertically(pages)
        cropped_img = crop_table_region(stitched_img)

        preview_images.append((uploaded_file.name, cropped_img))

        text = pytesseract.image_to_string(cropped_img, config="--psm 6")
        document_number = extract_document_number(text)

        lines = text.split("\n")
        capture = False
        current_colli = ""

        for line in lines:
            line = " ".join(line.split())
            line_lower = line.lower()

            colli_match = re.search(r"colli\s*#?\s*([0-9]+)", line, re.IGNORECASE)
            if colli_match:
                current_colli = colli_match.group(1).strip()

            if "item" in line_lower and "descr" in line_lower:
                capture = True
                continue

            if not capture:
                continue

            line = re.split(r"total\s+colli", line, flags=re.IGNORECASE)[0].strip()
            line_lower = line.lower()

            if line_lower.startswith("total colli"):
                capture = False
                continue

            if not line:
                continue

            item_match = re.search(r"^(\d{4}-\d{4})\s+(.+)$", line)
            if not item_match:
                continue

            item_no = item_match.group(1)
            remainder = item_match.group(2).strip()

            description = None
            quantity = None
            unit = None

            qty_match = re.search(
                r"(.+?)\s+([\d.,]+)\s+(piece|pcs?|bundle|kg|pc|roll|rolle|set|sets|meter|metre|Rolle|each)\b",
                remainder,
                re.IGNORECASE,
            )

            if qty_match:
                description = qty_match.group(1).strip()
                qty_raw = qty_match.group(2).strip()
                unit = qty_match.group(3).strip().lower()

                try:
                    quantity = normalize_quantity_text(qty_raw)
                except Exception:
                    quantity = None

            if quantity is None:
                fallback_match = re.search(
                    r"(.+?)\s+([\d.,]+)(?:\s+\w+)?$",
                    remainder,
                    re.IGNORECASE,
                )

                if fallback_match:
                    description = fallback_match.group(1).strip()
                    qty_raw = fallback_match.group(2).strip()
                    unit = "fallback"

                    try:
                        quantity = normalize_quantity_text(qty_raw)
                    except Exception:
                        quantity = None

            if quantity is None or description is None:
                continue

            all_items.append(
                {
                    "ItemNo": item_no,
                    "Description": description,
                    "Quantity": quantity,
                    "Unit": unit,
                    "ColliNo": current_colli,
                    "DocumentNumber": document_number,
                    "SourceFile": uploaded_file.name,
                    "PageNumber": "ALL",
                }
            )

    raw_df = pd.DataFrame(all_items)

    if raw_df.empty:
        summary_df = pd.DataFrame(columns=["ItemNo", "Description", "Quantity", "PalletList", "DocumentList"])
    else:
        summary_df = (
            raw_df.groupby("ItemNo")
            .agg(
                Quantity=("Quantity", "sum"),
                Description=("Description", lambda x: x.mode().iat[0] if not x.mode().empty else x.iloc[0]),
                PalletList=("ColliNo", combine_unique_values),
                DocumentList=("DocumentNumber", combine_unique_values),
            )
            .reset_index()
        )

        summary_df = summary_df[["ItemNo", "Description", "Quantity", "PalletList", "DocumentList"]]

    return raw_df, summary_df, preview_images


def build_tracker_lookup(wb):
    tracker_rows = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {}

        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is not None:
                headers[normalize_header(value)] = col

        item_col = headers.get("item #") or headers.get("part #")
        qty_col = headers.get("qty received")
        pallet_col = headers.get("pallet #")
        container_col = headers.get("container")

        if not item_col or not qty_col:
            continue

        for row in range(2, ws.max_row + 1):
            item_value = ws.cell(row=row, column=item_col).value
            if item_value is None:
                continue

            normalized_item = normalize_text(item_value)

            if normalized_item:
                if normalized_item not in tracker_rows:
                    tracker_rows[normalized_item] = []

                tracker_rows[normalized_item].append(
                    {
                        "sheet": ws,
                        "row": row,
                        "headers": headers,
                        "item_col": item_col,
                        "qty_col": qty_col,
                        "pallet_col": pallet_col,
                        "container_col": container_col,
                        "original_item": item_value,
                    }
                )

    return tracker_rows


def update_tracker_workbook(wb, summary_df, raw_df):
    tracker_rows = build_tracker_lookup(wb)

    matched = []
    not_found = []

    for _, row in summary_df.iterrows():
        raw_item_no = row["ItemNo"]
        item_no = normalize_part_number(raw_item_no)

        qty = row["Quantity"]
        desc = row["Description"]
        pallet_list = row["PalletList"]
        document_list = row["DocumentList"]

        if item_no in tracker_rows:
            entries = tracker_rows[item_no]

            for entry in entries:
                ws = entry["sheet"]
                excel_row = entry["row"]

                qty_col = entry["qty_col"]
                pallet_col = entry["pallet_col"]
                container_col = entry["container_col"]

                ws.cell(row=excel_row, column=qty_col).value = qty

                if pallet_col is not None:
                    existing_pallets = ws.cell(row=excel_row, column=pallet_col).value
                    merged_pallets = merge_comma_separated(existing_pallets, pallet_list)
                    ws.cell(row=excel_row, column=pallet_col).value = merged_pallets

                if container_col is not None:
                    existing_containers = ws.cell(row=excel_row, column=container_col).value
                    merged_containers = merge_comma_separated(existing_containers, document_list)
                    ws.cell(row=excel_row, column=container_col).value = merged_containers

                matched.append(
                    {
                        "ItemNo": item_no,
                        "Description": desc,
                        "Quantity": qty,
                        "PalletList": pallet_list,
                        "DocumentList": document_list,
                        "Sheet": ws.title,
                        "Row": excel_row,
                    }
                )
        else:
            not_found.append(
                {
                    "ItemNo": item_no,
                    "Description": desc,
                    "Quantity": qty,
                    "PalletList": pallet_list,
                    "DocumentList": document_list,
                }
            )

    unmatched_rows = []

    item_column = "Item #" if "Item #" in raw_df.columns else "ItemNo" if "ItemNo" in raw_df.columns else None
    if item_column is None:
        raise ValueError(f"Could not find item column in raw_df. Columns found: {list(raw_df.columns)}")

    for _, row in raw_df.iterrows():
        raw_item_no = row[item_column]
        item_no = normalize_part_number(raw_item_no)

        desc = row["Description"] if "Description" in raw_df.columns else ""
        qty = row["Quantity"] if "Quantity" in raw_df.columns else ""
        source_file = row["SourceFile"] if "SourceFile" in raw_df.columns else ""
        page_number = row["PageNumber"] if "PageNumber" in raw_df.columns else ""
        colli_no = row["ColliNo"] if "ColliNo" in raw_df.columns else ""
        document_number = row["DocumentNumber"] if "DocumentNumber" in raw_df.columns else ""

        if item_no not in tracker_rows:
            unmatched_rows.append(
                {
                    "Item #": raw_item_no,
                    "Normalized Item #": item_no,
                    "Description": desc,
                    "Quantity": qty,
                    "ColliNo": colli_no,
                    "DocumentNumber": document_number,
                    "SourceFile": source_file,
                    "PageNumber": page_number,
                }
            )

    unmatched_df = pd.DataFrame(unmatched_rows)

    if not unmatched_df.empty:
        unmatched_df = (
            unmatched_df.groupby(
                ["Item #", "Normalized Item #", "Description", "ColliNo", "DocumentNumber", "SourceFile", "PageNumber"],
                as_index=False,
            )["Quantity"].sum()
        )

    matched_df = pd.DataFrame(matched)
    not_found_df = pd.DataFrame(not_found)

    return wb, matched_df, not_found_df, unmatched_df


def workbook_to_bytes(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def dataframe_to_excel_bytes(sheets_dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()


# -----------------------------
# UI
# -----------------------------
delivery_files = st.file_uploader(
    "Upload delivery PDFs/images",
    type=["pdf", "png", "jpg", "jpeg", "tif", "tiff"],
    accept_multiple_files=True,
)

tracker_file = st.file_uploader(
    "Upload tracker workbook",
    type=["xlsx", "xlsm"],
)

run_button = st.button("Process Files", type="primary")

if run_button:
    if not delivery_files:
        st.error("Please upload at least one delivery PDF/image.")
        st.stop()

    if not tracker_file:
        st.error("Please upload the tracker workbook.")
        st.stop()

    with st.spinner("Running OCR and updating tracker..."):
        raw_df, summary_df, preview_images = process_delivery_files(delivery_files)

        tracker_bytes = tracker_file.getvalue()
        wb = load_workbook(BytesIO(tracker_bytes))

        wb, matched_df, not_found_df, unmatched_df = update_tracker_workbook(wb, summary_df, raw_df)

        updated_tracker_bytes = workbook_to_bytes(wb)

        ocr_results_bytes = dataframe_to_excel_bytes(
            {
                "Raw_OCR_Data": raw_df if not raw_df.empty else pd.DataFrame(),
                "Summarized_Totals": summary_df if not summary_df.empty else pd.DataFrame(),
            }
        )

        unmatched_export_bytes = None
        if unmatched_df is not None and not unmatched_df.empty:
            unmatched_export_bytes = dataframe_to_excel_bytes({"Unmatched_OCR_Items": unmatched_df})

    st.success("Processing complete.")

    if preview_images:
        with st.expander("Preview OCR Images", expanded=False):
            for name, img in preview_images:
                st.markdown(f"**{name}**")
                st.image(img, use_container_width=True)

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Summarized Totals")
        st.dataframe(summary_df, use_container_width=True)

    with col2:
        st.subheader("Matched Rows")
        st.dataframe(matched_df, use_container_width=True)

    st.subheader("Unmatched Items")
    if unmatched_df is not None and not unmatched_df.empty:
        st.dataframe(unmatched_df, use_container_width=True)
    else:
        st.info("No unmatched items.")

    st.download_button(
        label="Download Updated Tracker",
        data=updated_tracker_bytes,
        file_name="Updated_Tracking_File.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download OCR Results",
        data=ocr_results_bytes,
        file_name="OCR_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if unmatched_export_bytes is not None:
        st.download_button(
            label="Download Unmatched Items",
            data=unmatched_export_bytes,
            file_name="Unmatched_OCR_Items.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
